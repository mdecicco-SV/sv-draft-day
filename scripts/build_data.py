#!/usr/bin/env python3
"""
Phase 2 data builder for sv-draft-day.

Assembles the static board data the hub joins against live picks:
  public/data/composite.json  composite rankings (name-keyed)
  public/data/orgs.json       per-team draft tendencies + org-dev rank
  public/data/meta.json       build timestamp + source provenance

Team interest (teamintel.json) is NOT baked here — the browser fetches it live
from sv-teamintel for draft-day freshness.

Run:  python3 scripts/build_data.py
"""
import csv, json, os, datetime, urllib.request
from collections import Counter, defaultdict
import openpyxl

# MLBAM team id -> abbrev (mirrors lib/teams.js) for the static draft-order seed.
ID_TO_ABBR = {
    108:"LAA",109:"ARI",110:"BAL",111:"BOS",112:"CHC",113:"CIN",114:"CLE",115:"COL",
    116:"DET",117:"HOU",118:"KC",119:"LAD",120:"WSH",121:"NYM",133:"ATH",134:"PIT",
    135:"SD",136:"SEA",137:"SF",138:"STL",139:"TB",140:"TEX",141:"TOR",142:"MIN",
    143:"PHI",144:"ATL",145:"CHW",146:"MIA",147:"NYY",158:"MIL",
}
DRAFT_YEAR = 2026

NAME_TO_ABBR = {
    "arizona diamondbacks":"ARI","atlanta braves":"ATL","baltimore orioles":"BAL","boston red sox":"BOS",
    "chicago cubs":"CHC","chicago white sox":"CHW","cincinnati reds":"CIN","cleveland guardians":"CLE",
    "colorado rockies":"COL","detroit tigers":"DET","houston astros":"HOU","kansas city royals":"KC",
    "los angeles angels":"LAA","los angeles dodgers":"LAD","miami marlins":"MIA","milwaukee brewers":"MIL",
    "minnesota twins":"MIN","new york mets":"NYM","new york yankees":"NYY","philadelphia phillies":"PHI",
    "pittsburgh pirates":"PIT","san diego padres":"SD","san francisco giants":"SF","seattle mariners":"SEA",
    "st. louis cardinals":"STL","st louis cardinals":"STL","tampa bay rays":"TB","texas rangers":"TEX",
    "toronto blue jays":"TOR","washington nationals":"WSH",
}
def name_to_abbr(name):
    if not name: return None
    n = str(name).strip().lower()
    if n in NAME_TO_ABBR: return NAME_TO_ABBR[n]
    if "athletics" in n: return "ATH"
    return None

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "public", "data")
COMPOSITE_REPO = os.path.expanduser("~/Desktop/claude/draft-rankings-composite")
COMPOSITE = os.path.join(COMPOSITE_REPO, "outputs", "composite_latest.json")
RANK_HISTORY = os.path.join(COMPOSITE_REPO, "outputs", "rank_history.json")
ORGREVIEW_DIR = os.path.expanduser("~/Desktop/claude/sv-org-review")

# rank_history.json keys players by the composite repo's canon(norm(name)) — import the
# same functions so lookups can never drift from how that repo normalizes names.
import sys
sys.path.insert(0, COMPOSITE_REPO)
import build_composite as _bc

# org-review sheet abbrev -> live-feed abbrev (only the A's differ)
SHEET_TO_FEED = {"OAK": "ATH"}
TEAM_SHEETS = ["ARI","ATL","BAL","BOS","CHC","CHW","CIN","CLE","COL","DET","HOU","KC",
               "LAA","LAD","MIA","MIL","MIN","NYM","NYY","OAK","PHI","PIT","SD","SF",
               "SEA","STL","TB","TEX","TOR","WSH"]


def latest_orgreview():
    files = [f for f in os.listdir(ORGREVIEW_DIR) if f.startswith("Org.Review.2026.update_") and f.endswith(".xlsx")]
    # filenames are update_M-D-26 — sort by parsed date, newest last
    def keyf(f):
        d = f.replace("Org.Review.2026.update_", "").replace(".xlsx", "")
        m, day, _ = d.split("-")
        return (int(m), int(day))
    files.sort(key=keyf)
    return os.path.join(ORGREVIEW_DIR, files[-1]) if files else os.path.join(ORGREVIEW_DIR, "Org.Review.2026.xlsx")


def classify_level(school, age):
    """HS / JUCO / College from school name, falling back to age."""
    s = (school or "").lower()
    if "community college" in s or "junior college" in s or s.endswith(" cc") or s.endswith(" jc"):
        return "JUCO"
    hs_kw = ["high school", "academy", "prep", "christian", "catholic", "(hs)"]
    col_kw = ["university", "college", "state", "institute", " tech", "a&m", "polytech"]
    is_hs = (" hs" in (" " + s) or s.endswith(" hs") or any(k in s for k in hs_kw))
    is_col = any(k in s for k in col_kw)
    if is_hs and not is_col:
        return "HS"
    if is_col:
        return "College"
    try:
        a = float(age)
        if a and a <= 18.5:
            return "HS"
        if a and a >= 19.5:
            return "College"
    except (ValueError, TypeError):
        pass
    return "Unk"


def pos_group(pos):
    p = (pos or "").upper()
    if "RHP" in p or "LHP" in p or p in ("P", "SP", "RP", "PITCHER"):
        return "P"
    if p == "C":
        return "C"
    if p in ("1B", "2B", "3B", "SS", "IF", "INF"):
        return "IF"
    if p in ("OF", "CF", "LF", "RF"):
        return "OF"
    return "Other"


def _src_rank(v):
    """A per-source rank -> int, or None if the source didn't rank the player."""
    try:
        n = int(float(v))
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def _load_hist():
    return json.load(open(RANK_HISTORY, encoding="utf-8")) if os.path.exists(RANK_HISTORY) else None


def build_composite():
    """Composite board from composite_latest.json, plus ATH ranks and composite-rank
    trends from rank_history.json (ATH is hidden on the public share board by design —
    the internal draft-day board wants it)."""
    src = json.load(open(COMPOSITE, encoding="utf-8"))
    hist = _load_hist()
    ath_ranks = ((hist or {}).get("source_history", {}).get("ATH") or {}).get("ranks", {})
    hist_players = (hist or {}).get("players", {})
    latest = lambda vals: next((v for v in reversed(vals or []) if v is not None), None)
    rows = []
    for r in src["players"]:
        try:
            rank = int(float(r.get("Rank_Weighted") or 0))
        except (ValueError, TypeError):
            continue
        if not rank:
            continue
        pos = (r.get("Pos") or "").strip()
        school = (r.get("School") or "").strip()
        age = r.get("Age")
        age = "" if age in (None, "") else (str(int(float(age))) if float(age) == int(float(age)) else str(age))
        agent = (r.get("Agent") or "").strip()
        key = _bc.canon(_bc.norm(r.get("Name") or ""))
        hp = hist_players.get(key)
        row = {
            "rank": rank,
            "name": (r.get("Name") or "").strip(),
            "pos": pos,
            "posGroup": pos_group(pos),
            "level": classify_level(school, age),
            "school": school,
            "age": age,
            # advising agency (null when unknown/unlisted)
            "agent": agent if agent and agent.lower() != "unknown" else None,
            "sources": int(float(r.get("Num_Sources") or 0)),
            # per-source board ranks (null = not ranked by that source)
            "ba": _src_rank(r.get("BA_Rank")),
            "mlb": _src_rank(r.get("MLB_Rank")),
            "overslot": _src_rank(r.get("OverSlot_Rank")),
            "espn": _src_rank(r.get("ESPN_Rank")),
            "pbr": _src_rank(r.get("PBR_Rank")),
            "ath": _src_rank(latest(ath_ranks.get(key))),
        }
        # composite rank across ranking editions (aligned to meta.json's histDates)
        if hp and any(v is not None for v in hp["rw"]):
            row["trend"] = hp["rw"]
        rows.append(row)
    rows.sort(key=lambda x: x["rank"])
    return rows


def rank_meta():
    """Per-source depth + as-of date (ATH's come from rank history — it's not in the
    composite build) and the snapshot dates the players' trend arrays align to."""
    src = json.load(open(COMPOSITE, encoding="utf-8"))
    sources = {s: {"depth": d, "asOf": (src.get("as_of") or {}).get(s)}
               for s, d in (src.get("sources") or {}).items()}
    hist = _load_hist()
    ath = (hist or {}).get("source_history", {}).get("ATH")
    if ath and ath.get("dates"):
        last = len(ath["dates"]) - 1
        depth = sum(1 for v in ath["ranks"].values() if len(v) > last and v[last] is not None)
        sources["ATH"] = {"depth": depth, "asOf": ath["dates"][-1]}
    return {"sources": sources, "histDates": (hist or {}).get("dates", [])}


def bucket(school_class):
    if not school_class:
        return None
    c = str(school_class).upper()
    if c.startswith("HS"):
        return "HS"
    if c.startswith("JC"):
        return "JUCO"
    if c.startswith("4YR"):
        return "4YR"
    return None


def build_orgs(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ORG Rankings -> composite org-dev rank (column "COMP")
    org_rank = {}
    rk = wb["ORG Rankings"]
    rkrows = list(rk.iter_rows(values_only=True))
    hdr_i = next((i for i, row in enumerate(rkrows) if row and "Org" in [str(c) for c in row]), None)
    if hdr_i is not None:
        hdr = [str(c) if c is not None else "" for c in rkrows[hdr_i]]
        oi = hdr.index("Org")
        ci = hdr.index("COMP") if "COMP" in hdr else None
        for row in rkrows[hdr_i + 1:]:
            if not row or row[oi] is None:
                continue
            ab = str(row[oi]).strip()
            ab = SHEET_TO_FEED.get(ab, ab)
            if ci is not None and row[ci] is not None:
                try:
                    org_rank[ab] = round(float(row[ci]), 1)
                except (ValueError, TypeError):
                    pass

    orgs = {}
    for sheet in TEAM_SHEETS:
        feed_ab = SHEET_TO_FEED.get(sheet, sheet)
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # header row contains 'Year'
        hi = next((i for i, r in enumerate(rows) if r and "Year" in [str(c) for c in r]), None)
        if hi is None:
            continue
        hdr = [str(c) if c is not None else "" for c in rows[hi]]
        col = {name: hdr.index(name) for name in ["Year", "Round", "Pos", "Class", "Player", "School"] if name in hdr}

        pos_ct = Counter()
        cls_ct = Counter()
        early_pos = Counter()   # rounds 1-3 only
        recent_r1 = []
        for r in rows[hi + 1:]:
            if "Year" not in col or r[col["Year"]] is None:
                continue
            try:
                yr = int(float(r[col["Year"]]))
                rnd = int(float(r[col["Round"]])) if r[col.get("Round", -1)] is not None else None
            except (ValueError, TypeError):
                continue
            pos = (str(r[col["Pos"]]).strip() if "Pos" in col and r[col["Pos"]] else "")
            grp = "P" if pos == "P" else "H"
            pos_ct[grp] += 1
            b = bucket(r[col["Class"]]) if "Class" in col else None
            if b:
                cls_ct[b] += 1
            if rnd and rnd <= 3:
                early_pos[grp] += 1
            if rnd == 1 and yr >= 2023 and "Player" in col and r[col["Player"]]:
                recent_r1.append({
                    "year": yr,
                    "player": str(r[col["Player"]]).strip(),
                    "pos": pos,
                    "school": (str(r[col["School"]]).strip() if "School" in col and r[col["School"]] else ""),
                })

        total = sum(pos_ct.values())
        early = sum(early_pos.values())
        known = cls_ct["HS"] + cls_ct["4YR"] + cls_ct["JUCO"]
        recent_r1.sort(key=lambda x: -x["year"])
        orgs[feed_ab] = {
            "orgRank": org_rank.get(feed_ab),
            "n": total,
            "pitcherPct": round(100 * pos_ct["P"] / total) if total else None,
            "earlyPitcherPct": round(100 * early_pos["P"] / early) if early else None,
            "hsPct": round(100 * cls_ct["HS"] / known) if known else None,
            "collegePct": round(100 * (cls_ct["4YR"] + cls_ct["JUCO"]) / known) if known else None,
            "classN": known,
            "recentR1": recent_r1[:4],
        }

    # merge richer org-review metrics (development outcomes, draft-type trends, homegrown)
    extras = parse_claude_pull(wb)
    for ab, ex in extras.items():
        if ab in orgs:
            orgs[ab].update(ex)
    return orgs


def parse_claude_pull(wb):
    """Per-team development + draft-trend metrics from the 'Claude Data Pull' sheet
    (teams are columns). Returns { ABBR: {dev:{...}, trends:{...}, homegrown} }."""
    if "Claude Data Pull" not in wb.sheetnames:
        return {}
    rows = list(wb["Claude Data Pull"].iter_rows(values_only=True))
    hdr = next((r for r in rows if r and any(isinstance(c, str) and "Diamondbacks" in c for c in r)), None)
    if not hdr:
        return {}
    col_abbr = {i: name_to_abbr(c) for i, c in enumerate(hdr) if name_to_abbr(c)}
    def row_for(label):
        for r in rows:
            lab = next((str(x) for x in r[:2] if x not in (None, "")), "")
            if lab.strip().lower() == label.lower():
                return r
        return None
    def vals(label):
        r = row_for(label); out = {}
        if r:
            for i, ab in col_abbr.items():
                v = r[i] if i < len(r) else None
                if isinstance(v, (int, float)): out[ab] = v
        return out
    fields = {
        "mlbDebut": "MLB Debut %", "arb": "Reach Arbitration %",
        "posDebut": "Position Debut %", "pitDebut": "Pitcher Debut %",
        "ncaaPos": "NCAA Pos", "ncaaPitcher": "NCAA Pitcher", "hsPos": "HS Pos", "hsPitcher": "HS Pitcher",
        "homegrown": "Homegrown",
    }
    data = {k: vals(v) for k, v in fields.items()}
    out = {}
    for ab in col_abbr.values():
        out[ab] = {
            "dev": {k: round(data[k][ab], 3) for k in ("mlbDebut","arb","posDebut","pitDebut") if ab in data[k]},
            "trends": {k: round(data[k][ab], 3) for k in ("ncaaPos","ncaaPitcher","hsPos","hsPitcher") if ab in data[k]},
        }
        if ab in data["homegrown"]: out[ab]["homegrown"] = int(data["homegrown"][ab])
    return out


def _fetch_draft(year):
    url = f"https://statsapi.mlb.com/api/v1/draft/{year}"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    return json.loads(urllib.request.urlopen(req, timeout=30).read())


def _level_class(school_class):
    """HS / JUCO / College-JR / College-SR from the feed's schoolClass."""
    if not school_class:
        return None
    c = school_class.upper()
    if c.startswith("HS"):
        return "HS"
    if c.startswith("JC"):
        return "JUCO"
    if c.startswith("4YR"):
        return "College-SR" if ("SR" in c or "GR" in c or "5S" in c) else "College-JR"
    return None


# Leverage is strongly rank-dependent: top-ranked HS hitters sign at ~93% of their rank's
# slot while deep prep talent gets bought out of college commitments at 2-3x. One median
# per level x pos (plus a hand-tuned ramp) overshot the top of the board — so calibrate
# per rank band and let the front end interpolate between band midpoints.
LEV_BANDS = [(1, 9), (10, 20), (21, 45), (46, 120), (121, None)]
LEV_BAND_KEYS = [f"{lo}-{hi}" if hi else f"{lo}+" for lo, hi in LEV_BANDS]


def _lev_band(rank):
    for (lo, hi), key in zip(LEV_BANDS, LEV_BAND_KEYS):
        if rank >= lo and (hi is None or rank <= hi):
            return key


def build_leverage(years=(2021, 2022, 2023, 2024, 2025)):
    """Leverage multiplier = actual bonus / slot-at-pre-draft-rank, by level-class
    x rank band x position, calibrated from past drafts (statsapi embeds MLB's
    pre-draft rank). Thin cells shrink toward their parent (pos -> band -> level)."""
    import statistics
    ratios = defaultdict(list)
    for y in years:
        raw = _fetch_draft(y)
        slot_by_pick, slots = {}, []
        for rd in raw.get("drafts", {}).get("rounds", []):
            for p in rd.get("picks", []):
                sv = p.get("pickValue")
                if sv and float(sv) > 0 and p.get("pickNumber"):
                    slot_by_pick[int(p["pickNumber"])] = float(sv)
                    slots.append(float(sv))
        min_slot = min(slots) if slots else 150000.0
        for rd in raw.get("drafts", {}).get("rounds", []):
            for p in rd.get("picks", []):
                rank, bn = p.get("rank"), p.get("signingBonus")
                if not rank or not bn:
                    continue
                market = slot_by_pick.get(int(rank), min_slot)
                lv = _level_class((p.get("school") or {}).get("schoolClass"))
                pos = pos_group((p.get("person") or {}).get("primaryPosition", {}).get("abbreviation"))
                if lv and pos != "Other" and market > 0:
                    ratios[(lv, _lev_band(int(rank)), pos)].append(float(bn) / market)

    all_r = [x for arr in ratios.values() for x in arr]
    g = round(statistics.median(all_r), 3) if all_r else 1.0
    LEVELS = ["HS", "JUCO", "College-JR", "College-SR"]
    POS = ["P", "C", "IF", "OF"]
    out = {"mult": {"_": g}, "n": {},
           "meta": {"years": list(years), "global": g, "bands": LEV_BAND_KEYS}}
    for lv in LEVELS:
        lv_all = [x for key in LEV_BAND_KEYS for pos in POS for x in ratios.get((lv, key, pos), [])]
        lv_med = statistics.median(lv_all) if lv_all else g
        out["mult"][lv] = {"_": round(lv_med, 3)}
        out["n"][lv] = {}
        for key in LEV_BAND_KEYS:
            b_all = [x for pos in POS for x in ratios.get((lv, key, pos), [])]
            b_med = statistics.median(b_all) if b_all else lv_med
            if len(b_all) < 8:  # thin band -> shrink toward the level median
                b_med = (b_med * len(b_all) + lv_med * 8) / (len(b_all) + 8)
            cell, nn = {"_": round(b_med, 3)}, {"_": len(b_all)}
            for pos in POS:
                arr = ratios.get((lv, key, pos), [])
                n = len(arr)
                if n:
                    med = statistics.median(arr)
                    if n < 10:  # shrink small cells toward the band median
                        med = (med * n + b_med * 10) / (n + 10)
                    cell[pos] = round(med, 3)
                else:
                    cell[pos] = round(b_med, 3)
                nn[pos] = n
            out["mult"][lv][key] = cell
            out["n"][lv][key] = nn
    return out


def build_order():
    """Static draft-order skeleton (teams + slots, all undrafted) so the board
    renders on any server without the /api/draft serverless function."""
    url = f"https://statsapi.mlb.com/api/v1/draft/{DRAFT_YEAR}"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    raw = json.loads(urllib.request.urlopen(req, timeout=20).read())
    picks = []
    for rd in raw.get("drafts", {}).get("rounds", []):
        for p in rd.get("picks", []):
            tid = (p.get("team") or {}).get("id")
            sv = p.get("pickValue")
            picks.append({
                "overall": int(p["pickNumber"]) if p.get("pickNumber") else None,
                "round": str(p.get("pickRound")) if p.get("pickRound") is not None else None,
                "roundPick": int(p["roundPickNumber"]) if p.get("roundPickNumber") else None,
                "teamId": tid,
                "team": ID_TO_ABBR.get(tid),
                "teamName": (p.get("team") or {}).get("name"),
                "slot": float(sv) if sv not in (None, "") else None,
                "isDrafted": False, "isPass": False, "player": None, "playerId": None,
                "school": None, "schoolClass": None, "bucket": "UNK", "bonus": None, "rank": None,
            })
    picks.sort(key=lambda x: x["overall"] or 1e9)
    return {"year": DRAFT_YEAR, "picks": picks}


def main():
    os.makedirs(OUT, exist_ok=True)
    composite = build_composite()
    orgpath = latest_orgreview()
    orgs = build_orgs(orgpath)

    order = build_order()
    leverage = build_leverage()
    with open(os.path.join(OUT, "composite.json"), "w") as f:
        json.dump(composite, f, separators=(",", ":"))
    with open(os.path.join(OUT, "orgs.json"), "w") as f:
        json.dump(orgs, f, separators=(",", ":"))
    with open(os.path.join(OUT, "order.json"), "w") as f:
        json.dump(order, f, separators=(",", ":"))
    with open(os.path.join(OUT, "leverage.json"), "w") as f:
        json.dump(leverage, f, separators=(",", ":"))
    print(f"order.json      {len(order['picks'])} picks (static board seed)")
    print(f"leverage.json   HS={leverage['mult']['HS']}  CollegeJR={leverage['mult']['College-JR']}")
    with open(os.path.join(OUT, "meta.json"), "w") as f:
        json.dump({
            "builtAt": datetime.datetime.now().isoformat(timespec="seconds"),
            "composite_n": len(composite),
            "orgreview_file": os.path.basename(orgpath),
            **rank_meta(),
        }, f, indent=2)

    print(f"composite.json  {len(composite)} players")
    print(f"orgs.json       {len(orgs)} teams  (from {os.path.basename(orgpath)})")
    print(f"sample org KC:  {json.dumps(orgs.get('KC'))[:140]}")


if __name__ == "__main__":
    main()
